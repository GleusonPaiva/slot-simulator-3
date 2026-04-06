"""
=============================================================
  SLOT SIMULATOR 3 — Modo Bônus, Free Spins e Jackpot
=============================================================
  Projeto 3 de portfólio para Game Math Design em iGaming
  
  O que este projeto demonstra:
  - Jogo base com 5 rolos × 3 linhas e 10 paylines
  - Modo bônus acionado por 3+ Scatters
  - Free Spins com multiplicadores progressivos (2x a 10x)
  - Jackpot progressivo (Mini, Minor, Major, Grand)
  - RTP separado: jogo base vs bônus vs jackpot
  - Game Math Document nível certificação
=============================================================
"""

import random
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import os
import datetime

# ─────────────────────────────────────────────
#  1. CONFIGURAÇÃO DO JOGO
# ─────────────────────────────────────────────

NOME_DO_JOGO = "Dragon's Fortune"
VERSAO = "1.0.0"
NUM_ROLOS = 5
NUM_LINHAS_VISIVEIS = 3
NUM_LINHAS_PAGAMENTO = 10
APOSTA_POR_LINHA = 1.0
APOSTA_TOTAL = APOSTA_POR_LINHA * NUM_LINHAS_PAGAMENTO

# Símbolos — inclui Wild e Scatter
SIMBOLOS = ["Cereja", "Limao", "Laranja", "Estrela", "Diamante", "Sete", "Wild", "Scatter"]

# Pesos dos símbolos no JOGO BASE
PESOS_BASE = {
    "Cereja":   35,
    "Limao":    28,
    "Laranja":  18,
    "Estrela":  10,
    "Diamante":  5,
    "Sete":      2,
    "Wild":      1,   # Wild substitui qualquer símbolo
    "Scatter":   1,   # Scatter aciona o bônus
}

# Pesos dos símbolos no MODO BÔNUS (mais Wilds e Scatters)
PESOS_BONUS = {
    "Cereja":   30,
    "Limao":    25,
    "Laranja":  16,
    "Estrela":   9,
    "Diamante":  4,
    "Sete":      2,
    "Wild":      3,   # mais Wilds no bônus!
    "Scatter":   1,
}

# Tabela de pagamentos (Wild paga como Sete)
PAGAMENTOS = {
    "Cereja":   {3: 2,   4: 5,   5: 10},
    "Limao":    {3: 3,   4: 8,   5: 15},
    "Laranja":  {3: 5,   4: 12,  5: 25},
    "Estrela":  {3: 10,  4: 25,  5: 50},
    "Diamante": {3: 25,  4: 75,  5: 200},
    "Sete":     {3: 50,  4: 150, 5: 500},
    "Wild":     {3: 50,  4: 150, 5: 500},  # Wild paga igual ao Sete
}

# Configuração do BÔNUS
BONUS_CONFIG = {
    "scatters_para_ativar": 3,      # 3+ Scatters ativa o bônus
    "free_spins_por_scatter": {
        3: 10,  # 3 scatters = 10 free spins
        4: 15,  # 4 scatters = 15 free spins
        5: 25,  # 5 scatters = 25 free spins
    },
    "multiplicadores": [2, 3, 4, 5, 6, 8, 10],  # multiplicadores disponíveis
    "prob_aumentar_mult": 0.20,     # 20% de chance de aumentar o multiplicador a cada spin
    "retrigger_prob": 0.05,         # 5% de chance de reativar o bônus durante free spins
}

# Configuração do JACKPOT PROGRESSIVO
JACKPOT_CONFIG = {
    "contribuicao": 0.01,           # 1% de cada aposta vai para o jackpot
    "niveis": {
        "Mini":  {"prob": 0.005,  "valor_inicial": 10},
        "Minor": {"prob": 0.001,  "valor_inicial": 50},
        "Major": {"prob": 0.0002, "valor_inicial": 200},
        "Grand": {"prob": 0.00005,"valor_inicial": 1000},
    }
}

# Linhas de pagamento (mesmo do Projeto 2)
LINHAS_PAGAMENTO = {
    1:  [1, 1, 1, 1, 1],
    2:  [0, 0, 0, 0, 0],
    3:  [2, 2, 2, 2, 2],
    4:  [0, 1, 2, 1, 0],
    5:  [2, 1, 0, 1, 2],
    6:  [0, 0, 1, 2, 2],
    7:  [2, 2, 1, 0, 0],
    8:  [1, 0, 0, 0, 1],
    9:  [1, 2, 2, 2, 1],
    10: [0, 1, 0, 1, 0],
}

# ─────────────────────────────────────────────
#  2. MOTOR DO SLOT
# ─────────────────────────────────────────────

def criar_rolo(pesos):
    rolo = []
    for simbolo in SIMBOLOS:
        rolo.extend([simbolo] * pesos.get(simbolo, 0))
    return rolo

def girar_grade(rolos):
    grade = []
    for rolo in rolos:
        pos = random.randint(0, len(rolo) - 1)
        coluna = [
            rolo[(pos) % len(rolo)],
            rolo[(pos + 1) % len(rolo)],
            rolo[(pos + 2) % len(rolo)],
        ]
        grade.append(coluna)
    return grade

def contar_scatters(grade):
    """Conta quantos Scatters aparecem na grade (qualquer posição)."""
    total = 0
    for rolo in grade:
        for simbolo in rolo:
            if simbolo == "Scatter":
                total += 1
    return total

def aplicar_wild(simbolo):
    """Wild substitui qualquer símbolo exceto Scatter."""
    return simbolo  # Wild é tratado na lógica de linha

def verificar_linha_com_wild(grade, padrao_linha):
    """
    Verifica linha com suporte a Wild.
    Wild substitui qualquer símbolo na sequência.
    """
    simbolos_na_linha = [grade[rolo][padrao_linha[rolo]] for rolo in range(NUM_ROLOS)]
    
    # Determina o símbolo base (primeiro não-Wild)
    simbolo_base = None
    for s in simbolos_na_linha:
        if s != "Wild" and s != "Scatter":
            simbolo_base = s
            break
    
    if simbolo_base is None:
        simbolo_base = "Wild"  # linha toda de Wilds
    
    # Conta sequência da esquerda com Wild como coringa
    contagem = 0
    for s in simbolos_na_linha:
        if s == simbolo_base or s == "Wild":
            contagem += 1
        else:
            break
    
    return simbolo_base, contagem

def calcular_premio_grade(grade, multiplicador=1, aposta=APOSTA_POR_LINHA):
    """Calcula prêmio total da grade com multiplicador."""
    premio_total = 0.0
    
    for num_linha, padrao in LINHAS_PAGAMENTO.items():
        simbolo, contagem = verificar_linha_com_wild(grade, padrao)
        if contagem >= 3 and simbolo in PAGAMENTOS:
            mult_pag = PAGAMENTOS[simbolo].get(contagem, 0)
            if mult_pag > 0:
                premio = aposta * mult_pag * multiplicador
                premio_total += premio
    
    return premio_total

# ─────────────────────────────────────────────
#  3. MODO BÔNUS — FREE SPINS
# ─────────────────────────────────────────────

def executar_bonus(rolos_bonus):
    """
    Executa o modo bônus completo.
    Retorna prêmio total do bônus e estatísticas.
    """
    num_scatters = random.randint(3, 3)  # sempre 3 para simplificar
    num_free_spins = BONUS_CONFIG["free_spins_por_scatter"][num_scatters]
    
    multiplicador_atual = 1
    premio_bonus = 0.0
    spins_executados = 0
    retriggered = False

    while spins_executados < num_free_spins:
        # Verifica se aumenta o multiplicador
        if random.random() < BONUS_CONFIG["prob_aumentar_mult"]:
            mult_disponiveis = [m for m in BONUS_CONFIG["multiplicadores"] 
                               if m > multiplicador_atual]
            if mult_disponiveis:
                multiplicador_atual = mult_disponiveis[0]

        # Gira com os rolos do bônus
        grade = girar_grade(rolos_bonus)
        premio = calcular_premio_grade(grade, multiplicador=multiplicador_atual)
        premio_bonus += premio

        # Verifica retrigger
        scatters = contar_scatters(grade)
        if scatters >= 3 and not retriggered:
            num_free_spins += BONUS_CONFIG["free_spins_por_scatter"].get(scatters, 5)
            retriggered = True

        spins_executados += 1

    return premio_bonus, spins_executados, multiplicador_atual

# ─────────────────────────────────────────────
#  4. JACKPOT PROGRESSIVO
# ─────────────────────────────────────────────

class JackpotProgressivo:
    def __init__(self):
        self.valores = {
            nivel: config["valor_inicial"] 
            for nivel, config in JACKPOT_CONFIG["niveis"].items()
        }
        self.total_pago = 0.0
        self.hits = {nivel: 0 for nivel in JACKPOT_CONFIG["niveis"]}

    def contribuir(self, aposta):
        """Adiciona contribuição ao jackpot."""
        contribuicao = aposta * JACKPOT_CONFIG["contribuicao"]
        for nivel in self.valores:
            self.valores[nivel] += contribuicao * 0.25  # divide entre os 4 níveis

    def verificar_jackpot(self, aposta):
        """Verifica se ganhou algum nível de jackpot."""
        for nivel, config in JACKPOT_CONFIG["niveis"].items():
            if random.random() < config["prob"]:
                premio = self.valores[nivel]
                self.total_pago += premio
                self.hits[nivel] += 1
                self.valores[nivel] = config["valor_inicial"]  # reset
                return nivel, premio
        return None, 0

# ─────────────────────────────────────────────
#  5. SIMULAÇÃO MONTE CARLO
# ─────────────────────────────────────────────

def simular(num_rodadas=5_000_000, seed=42):
    """Simula o jogo completo com bônus e jackpot."""
    print(f"\n⚙️  Simulando {NOME_DO_JOGO}...")
    print(f"    {num_rodadas:,} rodadas...\n")

    random.seed(seed)
    rolos_base = [criar_rolo(PESOS_BASE) for _ in range(NUM_ROLOS)]
    rolos_bonus = [criar_rolo(PESOS_BONUS) for _ in range(NUM_ROLOS)]
    jackpot = JackpotProgressivo()

    # Contadores
    total_apostado = 0.0
    total_pago_base = 0.0
    total_pago_bonus = 0.0
    total_pago_jackpot = 0.0
    rodadas_bonus = 0
    historico_rtp = []
    checkpoint = num_rodadas // 10

    for i in range(num_rodadas):
        aposta = APOSTA_TOTAL
        total_apostado += aposta

        # Jackpot contribution
        jackpot.contribuir(aposta)

        # Gira o jogo base
        grade = girar_grade(rolos_base)
        premio_base = calcular_premio_grade(grade)
        total_pago_base += premio_base

        # Verifica Scatter para bônus
        num_scatters = contar_scatters(grade)
        if num_scatters >= BONUS_CONFIG["scatters_para_ativar"]:
            rodadas_bonus += 1
            premio_bonus, _, _ = executar_bonus(rolos_bonus)
            total_pago_bonus += premio_bonus

        # Verifica jackpot
        nivel_jackpot, premio_jackpot = jackpot.verificar_jackpot(aposta)
        if nivel_jackpot:
            total_pago_jackpot += premio_jackpot

        # Snapshot do RTP
        if (i + 1) % checkpoint == 0:
            total_pago = total_pago_base + total_pago_bonus + total_pago_jackpot
            rtp_atual = (total_pago / total_apostado) * 100
            historico_rtp.append({"rodada": i + 1, "rtp": rtp_atual})
            progresso = int((i + 1) / num_rodadas * 10)
            barra = "█" * progresso + "░" * (10 - progresso)
            print(f"    [{barra}] {(i+1)/num_rodadas*100:.0f}%  RTP: {rtp_atual:.2f}%")

    total_pago = total_pago_base + total_pago_bonus + total_pago_jackpot
    rtp_final = (total_pago / total_apostado) * 100
    rtp_base = (total_pago_base / total_apostado) * 100
    rtp_bonus = (total_pago_bonus / total_apostado) * 100
    rtp_jackpot = (total_pago_jackpot / total_apostado) * 100
    freq_bonus = (rodadas_bonus / num_rodadas) * 100

    print(f"\n✅ Simulação concluída!")
    print(f"   RTP Total:   {rtp_final:.4f}%")
    print(f"   RTP Base:    {rtp_base:.4f}%")
    print(f"   RTP Bônus:   {rtp_bonus:.4f}%")
    print(f"   RTP Jackpot: {rtp_jackpot:.4f}%")
    print(f"   Freq. Bônus: {freq_bonus:.4f}%")

    return {
        "rtp_final": rtp_final,
        "rtp_base": rtp_base,
        "rtp_bonus": rtp_bonus,
        "rtp_jackpot": rtp_jackpot,
        "freq_bonus": freq_bonus,
        "total_apostado": total_apostado,
        "total_pago": total_pago,
        "num_rodadas": num_rodadas,
        "rodadas_bonus": rodadas_bonus,
        "jackpot_hits": jackpot.hits,
        "historico_rtp": historico_rtp,
    }

# ─────────────────────────────────────────────
#  6. GRÁFICOS
# ─────────────────────────────────────────────

def gerar_graficos(resultados, output_dir):
    graficos = []
    plt.style.use('dark_background')
    BG = "#1a1a2e"
    BG2 = "#16213e"

    # ── Gráfico 1: Decomposição do RTP ──
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG2)

    componentes = ['Jogo Base', 'Bônus\n(Free Spins)', 'Jackpot', 'RTP Total']
    valores = [
        resultados["rtp_base"],
        resultados["rtp_bonus"],
        resultados["rtp_jackpot"],
        resultados["rtp_final"],
    ]
    cores = ['#3498db', '#9b59b6', '#f39c12', '#2ecc71']

    bars = ax.bar(componentes, valores, color=cores, edgecolor='#333', width=0.5)
    for bar, val in zip(bars, valores):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                f'{val:.2f}%', ha='center', color='white', fontweight='bold', fontsize=11)

    ax.set_ylabel('Contribuição RTP (%)', color='white')
    ax.set_title('Decomposição do RTP por Componente', color='white', fontsize=13, fontweight='bold')
    ax.tick_params(colors='white')
    ax.grid(axis='y', color='#333', linestyle='--', alpha=0.5)

    caminho = os.path.join(output_dir, 'grafico_rtp_decomposicao.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor=BG)
    plt.close()
    graficos.append(caminho)

    # ── Gráfico 2: Convergência do RTP ──
    fig, ax = plt.subplots(figsize=(11, 5))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG2)

    rodadas = [h["rodada"] for h in resultados["historico_rtp"]]
    rtps = [h["rtp"] for h in resultados["historico_rtp"]]

    ax.plot(rodadas, rtps, color='#00d4ff', linewidth=2, label='RTP Simulado')
    ax.axhline(y=resultados["rtp_final"], color='#ffd700', linestyle='--',
               linewidth=1.5, label=f'RTP Final ({resultados["rtp_final"]:.2f}%)')

    ax.set_xlabel('Número de Rodadas', color='white')
    ax.set_ylabel('RTP (%)', color='white')
    ax.set_title('Convergência do RTP — Dragon\'s Fortune', color='white', fontsize=13, fontweight='bold')
    ax.legend(facecolor=BG, edgecolor='#444', labelcolor='white')
    ax.tick_params(colors='white')
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.1f}M'))
    ax.grid(color='#333', linestyle='--', alpha=0.5)

    caminho = os.path.join(output_dir, 'grafico_convergencia.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor=BG)
    plt.close()
    graficos.append(caminho)

    # ── Gráfico 3: Jackpot hits ──
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG2)

    niveis = list(resultados["jackpot_hits"].keys())
    hits = list(resultados["jackpot_hits"].values())
    cores_jp = ['#95a5a6', '#3498db', '#9b59b6', '#f39c12']

    bars = ax.bar(niveis, hits, color=cores_jp, edgecolor='#333', width=0.5)
    for bar, hit in zip(bars, hits):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                str(hit), ha='center', color='white', fontweight='bold')

    ax.set_ylabel('Número de Hits', color='white')
    ax.set_title('Jackpot Progressivo — Frequência de Hits', color='white', fontsize=13, fontweight='bold')
    ax.tick_params(colors='white')
    ax.grid(axis='y', color='#333', linestyle='--', alpha=0.5)

    caminho = os.path.join(output_dir, 'grafico_jackpot.png')
    plt.tight_layout()
    plt.savefig(caminho, dpi=150, bbox_inches='tight', facecolor=BG)
    plt.close()
    graficos.append(caminho)

    return graficos

# ─────────────────────────────────────────────
#  7. EXPORTAR EXCEL
# ─────────────────────────────────────────────

def exportar_excel(resultados, output_dir):
    caminho = os.path.join(output_dir, f'{NOME_DO_JOGO.replace(" ","_")}_Math_Document.xlsx')
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    dark_fill = PatternFill("solid", fgColor="16213e")
    mid_fill = PatternFill("solid", fgColor="0f3460")
    header_font = Font(color="00d4ff", bold=True, size=11)
    title_font = Font(color="FFD700", bold=True, size=14)
    white_font = Font(color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin', color='444444'),
        right=Side(style='thin', color='444444'),
        top=Side(style='thin', color='444444'),
        bottom=Side(style='thin', color='444444')
    )

    def style_header(ws, row, cols):
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

    def style_row(ws, row, cols, alt=False):
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col)
            c.fill = mid_fill if alt else dark_fill
            c.font = white_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

    # ── Aba 1: Resumo ──
    ws1 = wb.active
    ws1.title = "📊 Resumo"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25

    ws1['A1'] = f"GAME MATH DOCUMENT — {NOME_DO_JOGO.upper()}"
    ws1['A1'].font = title_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:B1')

    dados = [
        ("Jogo", NOME_DO_JOGO),
        ("Versão", VERSAO),
        ("Data", datetime.date.today().strftime("%d/%m/%Y")),
        ("", ""),
        ("CONFIGURAÇÃO", ""),
        ("Rolos", f"{NUM_ROLOS}"),
        ("Linhas visíveis", f"{NUM_LINHAS_VISIVEIS}"),
        ("Linhas de pagamento", f"{NUM_LINHAS_PAGAMENTO}"),
        ("", ""),
        ("RTP BREAKDOWN", ""),
        ("RTP Total", f"{resultados['rtp_final']:.4f}%"),
        ("RTP Jogo Base", f"{resultados['rtp_base']:.4f}%"),
        ("RTP Bônus (Free Spins)", f"{resultados['rtp_bonus']:.4f}%"),
        ("RTP Jackpot", f"{resultados['rtp_jackpot']:.4f}%"),
        ("", ""),
        ("BÔNUS", ""),
        ("Frequência do Bônus", f"{resultados['freq_bonus']:.4f}%"),
        ("Rodadas de Bônus", f"{resultados['rodadas_bonus']:,}"),
        ("Scatters para ativar", f"{BONUS_CONFIG['scatters_para_ativar']}"),
        ("Free Spins (3 scatters)", f"{BONUS_CONFIG['free_spins_por_scatter'][3]}"),
        ("Prob. aumentar multiplicador", f"{BONUS_CONFIG['prob_aumentar_mult']*100:.0f}%"),
        ("", ""),
        ("JACKPOT", ""),
        ("Contribuição por aposta", f"{JACKPOT_CONFIG['contribuicao']*100:.0f}%"),
        ("Hits Mini", f"{resultados['jackpot_hits']['Mini']}"),
        ("Hits Minor", f"{resultados['jackpot_hits']['Minor']}"),
        ("Hits Major", f"{resultados['jackpot_hits']['Major']}"),
        ("Hits Grand", f"{resultados['jackpot_hits']['Grand']}"),
    ]

    for i, (label, valor) in enumerate(dados, start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=valor)
        style_row(ws1, i, 2, i % 2 == 0)
        if label in ("CONFIGURAÇÃO", "RTP BREAKDOWN", "BÔNUS", "JACKPOT"):
            ws1.cell(row=i, column=1).font = Font(color="FFD700", bold=True, size=11)
            ws1.merge_cells(f'A{i}:B{i}')

    # ── Aba 2: Tabela de Pagamentos ──
    ws2 = wb.create_sheet("💰 Pagamentos")
    ws2.sheet_view.showGridLines = False

    headers = ["Símbolo", "Peso Base", "Peso Bônus", "3 em linha", "4 em linha", "5 em linha"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers))

    for i, simbolo in enumerate(SIMBOLOS, start=2):
        pags = PAGAMENTOS.get(simbolo, {})
        row_data = [
            simbolo,
            PESOS_BASE.get(simbolo, 0),
            PESOS_BONUS.get(simbolo, 0),
            f'{pags.get(3, "-")}x' if pags.get(3) else "—",
            f'{pags.get(4, "-")}x' if pags.get(4) else "—",
            f'{pags.get(5, "-")}x' if pags.get(5) else "—",
        ]
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=col, value=val)
        style_row(ws2, i, len(headers), i % 2 == 0)

    for col in range(1, len(headers)+1):
        ws2.column_dimensions[chr(64+col)].width = 16

    # ── Aba 3: Jackpot ──
    ws3 = wb.create_sheet("🏆 Jackpot")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Nível", "Probabilidade", "Valor Inicial", "Hits", "Frequência (em rodadas)"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    for i, (nivel, config) in enumerate(JACKPOT_CONFIG["niveis"].items(), start=2):
        hits = resultados["jackpot_hits"][nivel]
        freq = resultados["num_rodadas"] / hits if hits > 0 else "N/A"
        row_data = [
            nivel,
            f'{config["prob"]*100:.4f}%',
            f'{config["valor_inicial"]} créditos',
            hits,
            f'1 em {freq:,.0f}' if isinstance(freq, float) else "N/A",
        ]
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=col, value=val)
        style_row(ws3, i, len(headers3), i % 2 == 0)

    for col in range(1, len(headers3)+1):
        ws3.column_dimensions[chr(64+col)].width = 22

    wb.save(caminho)
    print(f"\n📊 Excel salvo: {caminho}")
    return caminho

# ─────────────────────────────────────────────
#  8. EXPORTAR PDF
# ─────────────────────────────────────────────

def exportar_pdf(resultados, graficos, output_dir):
    caminho = os.path.join(output_dir, f'{NOME_DO_JOGO.replace(" ","_")}_Report.pdf')
    doc = SimpleDocTemplate(caminho, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    from reportlab.platypus import Image as RLImage
    BG = colors.HexColor('#1a1a2e')
    ACCENT = colors.HexColor('#00d4ff')
    GOLD = colors.HexColor('#FFD700')
    WHITE = colors.white
    ALT = colors.HexColor('#16213e')

    s_title = ParagraphStyle('t', fontSize=18, textColor=GOLD, alignment=TA_CENTER,
                              fontName='Helvetica-Bold', spaceAfter=6)
    s_h2 = ParagraphStyle('h2', fontSize=13, textColor=ACCENT, spaceBefore=12,
                           spaceAfter=6, fontName='Helvetica-Bold')
    s_body = ParagraphStyle('b', fontSize=10, textColor=WHITE, spaceAfter=4,
                             fontName='Helvetica')
    s_cap = ParagraphStyle('c', fontSize=9, textColor=colors.HexColor('#aaaaaa'),
                            alignment=TA_CENTER, spaceAfter=8)

    elementos = []
    elementos.append(Spacer(1, 0.5*cm))
    elementos.append(Paragraph(f"GAME MATH DOCUMENT — {NOME_DO_JOGO.upper()}", s_title))
    elementos.append(Paragraph(f"Versão {VERSAO} | Com Modo Bônus e Jackpot Progressivo", s_title))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", s_cap))
    elementos.append(Spacer(1, 0.5*cm))

    # RTP Breakdown
    elementos.append(Paragraph("1. RTP BREAKDOWN", s_h2))
    rtp_data = [
        ["Componente", "RTP", "% do Total"],
        ["Jogo Base", f'{resultados["rtp_base"]:.4f}%',
         f'{resultados["rtp_base"]/resultados["rtp_final"]*100:.1f}%'],
        ["Bônus (Free Spins)", f'{resultados["rtp_bonus"]:.4f}%',
         f'{resultados["rtp_bonus"]/resultados["rtp_final"]*100:.1f}%'],
        ["Jackpot Progressivo", f'{resultados["rtp_jackpot"]:.4f}%',
         f'{resultados["rtp_jackpot"]/resultados["rtp_final"]*100:.1f}%'],
        ["TOTAL", f'{resultados["rtp_final"]:.4f}%', "100%"],
    ]
    t = Table(rtp_data, colWidths=[6*cm, 5*cm, 5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f3460')),
        ('TEXTCOLOR', (0,0), (-1,0), ACCENT),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [BG, ALT]),
        ('TEXTCOLOR', (0,1), (-1,-1), WHITE),
        ('FONTNAME', (0,1), (0,-1), 'Helvetica'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,-1), (-1,-1), GOLD),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#444444')),
        ('ROWHEIGHT', (0,0), (-1,-1), 22),
    ]))
    elementos.append(t)
    elementos.append(Spacer(1, 0.5*cm))

    # Gráficos
    elementos.append(Paragraph("2. ANÁLISE GRÁFICA", s_h2))
    captions = [
        "Figura 1 — Decomposição do RTP: jogo base, bônus e jackpot",
        "Figura 2 — Convergência do RTP ao longo das simulações",
        "Figura 3 — Frequência de hits por nível de jackpot",
    ]
    for i, grafico in enumerate(graficos):
        if os.path.exists(grafico):
            img = RLImage(grafico, width=16*cm, height=8*cm)
            elementos.append(img)
            elementos.append(Paragraph(captions[i] if i < len(captions) else "", s_cap))
            elementos.append(Spacer(1, 0.3*cm))

    # Configuração do Bônus
    elementos.append(Paragraph("3. CONFIGURAÇÃO DO MODO BÔNUS", s_h2))
    bonus_texto = f"""
    O modo bônus é ativado quando <b>3 ou mais Scatters</b> aparecem em qualquer posição da grade.
    Com 3 Scatters o jogador recebe <b>{BONUS_CONFIG["free_spins_por_scatter"][3]} Free Spins</b>,
    com 4 recebe <b>{BONUS_CONFIG["free_spins_por_scatter"][4]}</b> e com 5 recebe
    <b>{BONUS_CONFIG["free_spins_por_scatter"][5]}</b>.
    Durante os Free Spins, há <b>{BONUS_CONFIG["prob_aumentar_mult"]*100:.0f}%</b> de chance
    a cada spin de aumentar o multiplicador, que pode chegar até <b>10x</b>.
    A frequência de ativação do bônus foi de <b>{resultados["freq_bonus"]:.4f}%</b>
    ({resultados["rodadas_bonus"]:,} bônus em {resultados["num_rodadas"]:,} rodadas).
    """
    elementos.append(Paragraph(bonus_texto, s_body))

    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph("─" * 80, s_cap))
    elementos.append(Paragraph(f"Game Math Document — {NOME_DO_JOGO} v{VERSAO} — Portfólio iGaming", s_cap))

    def background(canvas, doc):
        canvas.setFillColor(BG)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)

    doc.build(elementos, onFirstPage=background, onLaterPages=background)
    print(f"📄 PDF salvo: {caminho}")
    return caminho

# ─────────────────────────────────────────────
#  9. MAIN
# ─────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print(f"  🐉 {NOME_DO_JOGO} — SLOT SIMULATOR 3")
    print("="*60)
    print("  Features: Wild · Scatter · Free Spins · Jackpot Progressivo")
    print("="*60)

    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    resultados = simular(num_rodadas=5_000_000)

    print("\n🎨 Gerando gráficos...")
    graficos = gerar_graficos(resultados, output_dir)

    print("\n📊 Gerando Excel...")
    exportar_excel(resultados, output_dir)

    print("\n📄 Gerando PDF...")
    exportar_pdf(resultados, graficos, output_dir)

    print("\n" + "="*60)
    print("  ✅ PROJETO 3 CONCLUÍDO!")
    print("="*60)
    print(f"\n  RTP Total:          {resultados['rtp_final']:.4f}%")
    print(f"  ├─ Jogo Base:       {resultados['rtp_base']:.4f}%")
    print(f"  ├─ Bônus (Spins):   {resultados['rtp_bonus']:.4f}%")
    print(f"  └─ Jackpot:         {resultados['rtp_jackpot']:.4f}%")
    print(f"\n  Frequência Bônus:   {resultados['freq_bonus']:.4f}%")
    print(f"  Rodadas de Bônus:   {resultados['rodadas_bonus']:,}")
    print(f"\n  Jackpot Hits:")
    for nivel, hits in resultados['jackpot_hits'].items():
        print(f"  ├─ {nivel}: {hits}")
    print(f"\n  Arquivos em: {output_dir}/")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()
